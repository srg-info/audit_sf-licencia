''' ======================================================
--  Proyecto: auditoria_soft_powershell
--  Autor: [serg.info]
--  Licencia: [MIT]
--  Descripción: Aquí te dejo el Script en Python para ejecuta desde la powersell.
-- ======================================================'''

import subprocess
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ================================
# CONFIGURACIÓN
# ================================
LISTA_PERMITIDA = "lista_software_permitido.txt"
REPORTE_EXCEL = "reporte_auditoria.xlsx"

verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Autorizado
rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # No autorizado

# ================================
# 1. CARGAR LISTA BLANCA
# ================================
try:
    with open(LISTA_PERMITIDA, "r", encoding="utf-8") as f:
        lista_permitida = [line.strip().lower() for line in f if line.strip()]
    print(f"✅ Lista blanca cargada ({len(lista_permitida)} programas permitidos)")
except FileNotFoundError:
    print(f"⚠️ No se encontró '{LISTA_PERMITIDA}', se considerará que todo es No autorizado.")
    lista_permitida = []

# ================================
# 2. OBTENER SOFTWARE INSTALADO (PowerShell)
# ================================
print("\n🔍 Escaneando software instalado (PowerShell)...")

comando = [
    "powershell",
    "-Command",
    "Get-WmiObject -Class Win32_Product | Select-Object Name, Version"
]

try:
    resultado = subprocess.check_output(comando, shell=True)
    lineas = resultado.decode("utf-8", errors="ignore").split("\n")
except Exception as e:
    print("❌ Error al obtener la lista de software:", e)
    exit()

software_detectado = []

for linea in lineas:
    if linea.strip() and not linea.strip().startswith("Name"):
        partes = [p.strip() for p in linea.split() if p.strip()]
        if len(partes) >= 1:
            nombre = " ".join(partes[:-1]).lower() if len(partes) > 1 else partes[0].lower()
            version = partes[-1] if len(partes) > 1 else "Desconocida"
            estado = "Autorizado" if any(p in nombre for p in lista_permitida) else "No autorizado"
            software_detectado.append([nombre.title(), version, estado])

print(f"✅ {len(software_detectado)} programas detectados.")

# ================================
# 3. GENERAR REPORTE EN EXCEL
# ================================
wb = Workbook()
ws = wb.active
ws.title = "Auditoría de Software"

# Encabezados
ws.append(["Nombre", "Versión", "Estado"])

# Datos con colores
for nombre, version, estado in software_detectado:
    ws.append([nombre, version, estado])
    color = verde if estado == "Autorizado" else rojo
    for col in range(1, 4):
        ws.cell(ws.max_row, col).fill = color

wb.save(REPORTE_EXCEL)
print(f"✅ Reporte generado: {REPORTE_EXCEL}")
