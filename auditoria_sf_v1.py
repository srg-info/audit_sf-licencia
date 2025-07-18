''' ======================================================
--  Proyecto: auditoria_sf_v1
--  Autor: [serg.info]
--  Licencia: [MIT]
--  Descripción: Aquí te dejo el Script en Python, realizado para el escaneo automático de los quipos.
--  Su función es detectar software instalado no licenciado o no oficial, generando un informe en Excel o CSV 
--  para la toma de decisiones en áreas de TI y seguridad, en su función arroja software absoleto en la lista no liceciados."
-- ======================================================'''

import subprocess
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# ================================
# CONFIGURACIÓN
# ================================
LISTA_PERMITIDA = "lista_software_permitido.txt"
REPORTE_COMPLETO = "reporte_auditoria.xlsx"
REPORTE_NO_AUTORIZADOS = "reporte_no_autorizados.xlsx"

verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Autorizado
rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # No autorizado

# ================================
# 1. CARGAR LISTA BLANCA
# ================================
try:
    with open(LISTA_PERMITIDA, "r", encoding="utf-8") as f:
        lista_permitida = [line.strip().lower() for line in f if line.strip()]
    print(f"✅ Lista blanca cargada ({len(lista_permitida)} programas permitidos)")
except FileNotFoundError:
    print(f"⚠️ No se encontró '{LISTA_PERMITIDA}', se considerará que todo es No autorizado.")
    lista_permitida = []

# ================================
# 2. OBTENER SOFTWARE INSTALADO (PowerShell)
# ================================
print("\n🔍 Escaneando software instalado...")

comando = [
    "powershell",
    "-Command",
    "Get-WmiObject -Class Win32_Product | Select-Object Name, Version"
]

try:
    resultado = subprocess.check_output(comando, shell=True)
    lineas = resultado.decode("utf-8", errors="ignore").split("\n")
except Exception as e:
    print("❌ Error al obtener la lista de software:", e)
    exit()

software_detectado = []

for linea in lineas:
    if linea.strip() and not linea.strip().startswith("Name"):
        partes = [p.strip() for p in linea.split() if p.strip()]
        if len(partes) >= 1:
            nombre = " ".join(partes[:-1]).lower() if len(partes) > 1 else partes[0].lower()
            version = partes[-1] if len(partes) > 1 else "Desconocida"
            estado = "Autorizado" if any(p in nombre for p in lista_permitida) else "No autorizado"
            software_detectado.append([nombre.title(), version, estado])

print(f"✅ {len(software_detectado)} programas detectados.")

# ================================
# 3. GENERAR REPORTE COMPLETO
# ================================
wb = Workbook()
ws = wb.active
ws.title = "Auditoría de Software"
ws.append(["Nombre", "Versión", "Estado"])

for nombre, version, estado in software_detectado:
    ws.append([nombre, version, estado])
    color = verde if estado == "Autorizado" else rojo
    for col in range(1, 3+1):
        ws.cell(ws.max_row, col).fill = color

wb.save(REPORTE_COMPLETO)
print(f"✅ Reporte completo generado: {REPORTE_COMPLETO}")

# ================================
# 4. GENERAR REPORTE SOLO NO AUTORIZADOS
# ================================
wb_na = Workbook()
ws_na = wb_na.active
ws_na.title = "No Autorizados"
ws_na.append(["Nombre", "Versión", "Estado"])

no_autorizados = [s for s in software_detectado if s[2] == "No autorizado"]

for nombre, version, estado in no_autorizados:
    ws_na.append([nombre, version, estado])
    for col in range(1, 3+1):
        ws_na.cell(ws_na.max_row, col).fill = rojo

wb_na.save(REPORTE_NO_AUTORIZADOS)
print(f"✅ Reporte de No Autorizados generado: {REPORTE_NO_AUTORIZADOS}")

# ================================
# 5. ABRIR AUTOMÁTICAMENTE LOS REPORTES
# ================================
try:
    os.startfile(REPORTE_COMPLETO)
    os.startfile(REPORTE_NO_AUTORIZADOS)
except:
    print("⚠️ No se pudieron abrir los archivos automáticamente, ábrelos manualmente.")
